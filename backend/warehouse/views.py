import csv
import json
from decimal import Decimal, InvalidOperation
from io import TextIOWrapper

from django.db.models import Case, DecimalField, F, Q, Sum, Value, When
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook

from .models import InventoryRecord, WarehouseItem


def record_to_dict(record):
    return {
        "id": record.id,
        "school": record.school,
        "schoolLabel": record.get_school_display(),
        "recordType": record.record_type,
        "recordTypeLabel": record.get_record_type_display(),
        "itemId": record.item_id,
        "itemName": record.item_name,
        "unitPrice": float(record.item.unit_price) if record.item else 0,
        "quantity": float(record.quantity),
        "unit": record.unit,
        "supplier": record.supplier,
        "operator": record.operator,
        "occurredAt": record.occurred_at.isoformat(),
        "remark": record.remark,
        "createdAt": record.created_at.isoformat(),
    }


def parse_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return {}


def filtered_records(request):
    qs = InventoryRecord.objects.all()
    school = request.GET.get("school")
    record_type = request.GET.get("recordType")
    keyword = request.GET.get("keyword")
    start_date = request.GET.get("startDate")
    end_date = request.GET.get("endDate")

    if school:
        qs = qs.filter(school=school)
    if record_type:
        qs = qs.filter(record_type=record_type)
    if start_date:
        qs = qs.filter(occurred_at__gte=start_date)
    if end_date:
        qs = qs.filter(occurred_at__lte=end_date)
    if keyword:
        qs = qs.filter(
            Q(item_name__icontains=keyword)
            | Q(supplier__icontains=keyword)
            | Q(operator__icontains=keyword)
        )
    return qs


def item_to_dict(item):
    return {
        "id": item.id,
        "name": item.name,
        "unit": item.unit,
        "unitPrice": float(item.unit_price),
        "isActive": item.is_active,
    }


def find_choice(value, choices):
    labels = {label: key for key, label in choices}
    if value in labels:
        return labels[value]
    if value in dict(choices):
        return value
    return None


def normalize_header(value):
    return str(value or "").strip().replace(" ", "")


def get_row_value(row, headers, names):
    for name in names:
        index = headers.get(name)
        if index is not None:
            value = row[index]
            if value is not None and str(value).strip() != "":
                return str(value).strip()
    return ""


def get_row_raw(row, headers, names):
    for name in names:
        index = headers.get(name)
        if index is not None and row[index] not in (None, ""):
            return row[index]
    return ""


def parse_import_date(value):
    if hasattr(value, "date"):
        return value.date()
    return parse_date(str(value).strip())


def create_record_from_data(data):
    item = None
    if data.get("itemId"):
        try:
            item = WarehouseItem.objects.get(pk=data["itemId"], is_active=True)
        except WarehouseItem.DoesNotExist:
            raise ValueError("物品不存在或未启用")
    else:
        item_name = str(data.get("itemName", "")).strip()
        if not item_name:
            raise ValueError("缺少物品名称")
        unit_price = Decimal(str(data.get("unitPrice") or 0))
        item, _ = WarehouseItem.objects.get_or_create(
            name=item_name,
            defaults={
                "unit": str(data.get("unit") or "kg").strip() or "kg",
                "unit_price": unit_price,
            },
        )
        if unit_price > 0 and item.unit_price != unit_price:
            item.unit_price = unit_price
            item.save(update_fields=["unit_price"])

    occurred_at = parse_import_date(data.get("occurredAt"))
    if occurred_at is None:
        raise ValueError("日期格式不正确")

    quantity = Decimal(str(data.get("quantity")))
    if quantity <= 0:
        raise ValueError("数量必须大于 0")

    unit = str(data.get("unit") or item.unit).strip() or item.unit
    record = InventoryRecord.objects.create(
        school=data["school"],
        record_type=data["recordType"],
        item=item,
        item_name=item.name,
        quantity=quantity,
        unit=unit,
        supplier=str(data.get("supplier", "")).strip(),
        operator=str(data.get("operator", "")).strip(),
        occurred_at=occurred_at,
        remark=str(data.get("remark", "")).strip(),
    )
    return record


@csrf_exempt
@require_http_methods(["GET", "POST"])
def items(request):
    if request.method == "GET":
        qs = WarehouseItem.objects.filter(is_active=True)
        return JsonResponse({"results": [item_to_dict(item) for item in qs]})

    data = parse_body(request)
    name = str(data.get("name", "")).strip()
    unit = str(data.get("unit", "kg")).strip() or "kg"
    try:
        unit_price = Decimal(str(data.get("unitPrice") or 0))
    except InvalidOperation:
        return JsonResponse({"error": "单价格式不正确"}, status=400)
    if not name:
        return JsonResponse({"error": "缺少物品名称"}, status=400)
    item, created = WarehouseItem.objects.get_or_create(
        name=name,
        defaults={"unit": unit, "unit_price": unit_price},
    )
    if not item.is_active:
        item.is_active = True
    item.unit = unit
    item.unit_price = unit_price
    item.save(update_fields=["is_active", "unit", "unit_price"])
    return JsonResponse(item_to_dict(item), status=201 if created else 200)


@csrf_exempt
@require_http_methods(["PATCH", "PUT", "DELETE"])
def item_detail(request, item_id):
    try:
        item = WarehouseItem.objects.get(pk=item_id, is_active=True)
    except WarehouseItem.DoesNotExist:
        return JsonResponse({"error": "物品不存在"}, status=404)

    if request.method == "DELETE":
        item.is_active = False
        item.save(update_fields=["is_active"])
        return JsonResponse({"ok": True})

    data = parse_body(request)
    name = str(data.get("name", item.name)).strip()
    unit = str(data.get("unit", item.unit)).strip() or item.unit
    try:
        unit_price = Decimal(str(data.get("unitPrice", item.unit_price) or 0))
    except InvalidOperation:
        return JsonResponse({"error": "单价格式不正确"}, status=400)
    if not name:
        return JsonResponse({"error": "缺少物品名称"}, status=400)

    item.name = name
    item.unit = unit
    item.unit_price = unit_price
    try:
        item.save(update_fields=["name", "unit", "unit_price"])
    except IntegrityError:
        return JsonResponse({"error": "物品名称已存在"}, status=400)

    InventoryRecord.objects.filter(item=item).update(item_name=item.name)
    return JsonResponse(item_to_dict(item))


@csrf_exempt
@require_http_methods(["GET", "POST"])
def records(request):
    if request.method == "GET":
        qs = filtered_records(request)
        return JsonResponse({"results": [record_to_dict(record) for record in qs]})

    data = parse_body(request)
    required = ["school", "recordType", "itemId", "quantity", "occurredAt"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return JsonResponse({"error": f"缺少字段: {', '.join(missing)}"}, status=400)

    try:
        record = create_record_from_data(data)
        return JsonResponse(record_to_dict(record), status=201)
    except (ValueError, KeyError, InvalidOperation) as error:
        return JsonResponse({"error": str(error)}, status=400)


@csrf_exempt
@require_http_methods(["DELETE"])
def record_detail(request, record_id):
    try:
        record = InventoryRecord.objects.get(pk=record_id)
    except InventoryRecord.DoesNotExist:
        return JsonResponse({"error": "记录不存在"}, status=404)
    record.delete()
    return JsonResponse({"ok": True})


@require_http_methods(["GET"])
def summary(request):
    qs = filtered_records(request)
    signed_quantity = Case(
        When(record_type=InventoryRecord.TYPE_IN, then=F("quantity")),
        When(record_type=InventoryRecord.TYPE_OUT, then=-F("quantity")),
        default=Value(0),
        output_field=DecimalField(max_digits=10, decimal_places=2),
    )
    totals = qs.aggregate(
        inQuantity=Sum("quantity", filter=Q(record_type=InventoryRecord.TYPE_IN)),
        outQuantity=Sum("quantity", filter=Q(record_type=InventoryRecord.TYPE_OUT)),
        stock=Sum(signed_quantity),
    )
    items = (
        qs.values("school", "item_name", "unit")
        .annotate(stock=Sum(signed_quantity))
        .order_by("school", "item_name")
    )
    return JsonResponse(
        {
            "inQuantity": float(totals["inQuantity"] or 0),
            "outQuantity": float(totals["outQuantity"] or 0),
            "stock": float(totals["stock"] or 0),
            "items": [
                {
                    "school": item["school"],
                    "itemName": item["item_name"],
                    "unit": item["unit"],
                    "stock": float(item["stock"] or 0),
                }
                for item in items
            ],
        }
    )


@require_http_methods(["GET"])
def export_records(request):
    file_type = request.GET.get("type", "xlsx")
    qs = filtered_records(request)
    rows = [
        [
            record.get_school_display(),
            record.get_record_type_display(),
            record.item_name,
            float(record.item.unit_price) if record.item else 0,
            float(record.quantity),
            record.unit,
            record.supplier,
            record.operator,
            record.occurred_at.isoformat(),
            record.remark,
        ]
        for record in qs
    ]
    headers = ["学校", "类型", "物品名称", "单价", "数量", "单位", "供应商/领用人", "经办人", "日期", "备注"]

    if file_type == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="warehouse-records.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "出入库记录"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 4
        sheet.column_dimensions[column[0].column_letter].width = min(width, 24)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="warehouse-records.xlsx"'
    workbook.save(response)
    return response


def iter_csv_rows(uploaded_file):
    wrapper = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
    yield from csv.reader(wrapper)


def iter_xlsx_rows(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        yield list(row)


@csrf_exempt
@require_http_methods(["POST"])
def import_records(request):
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"error": "请选择要导入的文件"}, status=400)

    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        rows = iter_csv_rows(uploaded_file)
    elif filename.endswith(".xlsx"):
        rows = iter_xlsx_rows(uploaded_file)
    else:
        return JsonResponse({"error": "仅支持 CSV 或 XLSX 文件"}, status=400)
    rows = list(rows)
    if not rows:
        return JsonResponse({"error": "文件为空"}, status=400)

    headers = {normalize_header(value): index for index, value in enumerate(rows[0])}
    created_count = 0
    errors = []
    for line_number, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        try:
            school = find_choice(get_row_value(row, headers, ["学校", "学校类型"]), InventoryRecord.SCHOOL_CHOICES)
            record_type = find_choice(get_row_value(row, headers, ["类型", "记录类型", "出入库"]), InventoryRecord.TYPE_CHOICES)
            data = {
                "school": school,
                "recordType": record_type,
                "itemName": get_row_value(row, headers, ["物品名称", "物品"]),
                "unitPrice": get_row_value(row, headers, ["单价", "价格"]),
                "quantity": get_row_value(row, headers, ["数量"]),
                "unit": get_row_value(row, headers, ["单位"]) or "kg",
                "supplier": get_row_value(row, headers, ["供应商/领用人", "供应商", "领用人"]),
                "operator": get_row_value(row, headers, ["经办人", "操作人"]),
                "occurredAt": get_row_raw(row, headers, ["日期", "发生日期"]),
                "remark": get_row_value(row, headers, ["备注"]),
            }
            if not data["school"]:
                raise ValueError("学校必须是 小学 或 中学")
            if not data["recordType"]:
                raise ValueError("类型必须是 入库 或 出库")
            create_record_from_data(data)
            created_count += 1
        except Exception as error:
            errors.append(f"第 {line_number} 行：{error}")

    status = 200 if not errors else 400
    return JsonResponse({"created": created_count, "errors": errors[:20]}, status=status)
